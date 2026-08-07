"""SHFE Gold 2008-2026 Full Backtest"""
import openpyxl

wb = openpyxl.load_workbook(r'../.reasonix/attachments/clipboard-20260806-110936.630003-000010.xlsx', data_only=True)
ws = wb.active

prices = []
for i in range(2, ws.max_row + 1):
    dt = ws.cell(i, 3).value
    close = ws.cell(i, 7).value
    if dt and close:
        prices.append((dt.strftime('%Y-%m-%d') if hasattr(dt, 'strftime') else str(dt)[:10], float(close)))

prices.sort(key=lambda x: x[0])
n = len(prices)
print(f"SHFE Gold: {prices[0][0]} ~ {prices[-1][0]}, {n} rows")
print(f"Price range: {prices[0][1]:.0f} ~ {prices[-1][1]:.0f} CNY/g")
print()


def ma_at(idx, period):
    if idx < period - 1:
        return None
    return sum(p[1] for p in prices[idx - period + 1:idx + 1]) / period


# ============================================
# Part 1: All Death Cross periods
# ============================================
print("=" * 115)
print("  ALL MA50/MA200 Death Cross Periods (SHFE Gold 2008-2026)")
print("=" * 115)

dc_periods = []
in_dc = False
dc_start = 0

for idx in range(200, n):
    ma50 = ma_at(idx, 50)
    ma200 = ma_at(idx, 200)
    is_dc = ma50 and ma200 and ma50 < ma200

    if is_dc and not in_dc:
        dc_start = idx
        in_dc = True
    elif not is_dc and in_dc:
        dc_end = idx - 1
        in_dc = False
        dc_days = dc_end - dc_start + 1
        ma20_end = ma_at(dc_end, 20)
        above_ma20_end = prices[dc_end][1] > ma20_end if ma20_end else False
        y1_high = max(p[1] for p in prices[max(0, dc_end - 252):dc_end + 1])
        dd_end = (prices[dc_end][1] / y1_high - 1) * 100
        fwd_rets = {}
        for m in [1, 3, 6, 12]:
            fwd_i = min(dc_end + 21 * m, n - 1)
            fwd_rets[m] = round((prices[fwd_i][1] / prices[dc_end][1] - 1) * 100, 1)
        dc_periods.append({
            'start': prices[dc_start][0], 'end': prices[dc_end][0],
            'days': dc_days,
            'start_price': round(prices[dc_start][1], 0),
            'end_price': round(prices[dc_end][1], 0),
            'low_price': round(min(p[1] for p in prices[dc_start:dc_end + 1]), 0),
            'dd_end': round(dd_end, 1), 'above_ma20': above_ma20_end,
            'fwd': fwd_rets,
        })

print(f"{'Period':<26} {'Days':>5} {'Price Journey':>24} {'DD%':>7} {'MA20':>6} {'1M':>7} {'3M':>7} {'6M':>7} {'12M':>8} {'Outcome':>16}")
print("-" * 115)

for dp in dc_periods:
    pr = f"{dp['start_price']:,.0f}>{dp['low_price']:,.0f}>{dp['end_price']:,.0f}"
    ma20_label = "ABOVE" if dp['above_ma20'] else "below"
    f1 = f"{dp['fwd'][1]:+.1f}%"
    f3 = f"{dp['fwd'][3]:+.1f}%"
    f6 = f"{dp['fwd'][6]:+.1f}%"
    f12 = f"{dp['fwd'][12]:+.1f}%"
    if dp['fwd'][6] > 10:
        result = "Bull resumed +++"
    elif dp['fwd'][6] > 0:
        result = "Slow recovery"
    elif dp['fwd'][6] > -10:
        result = "Sideways / down"
    else:
        result = "Bear continued ---"
    marker = " <-- SIMILAR NOW" if dp['end'] >= '2026-07' else ""
    print(f"{dp['start'][:10]}~{dp['end'][:10]} {dp['days']:>4}d {pr:>24} {dp['dd_end']:>+6.1f}% {ma20_label:>6} {f1:>7} {f3:>7} {f6:>7} {f12:>8} {result:>16}{marker}")

if in_dc:
    dc_end = n - 1
    dc_days = dc_end - dc_start + 1
    y1_high = max(p[1] for p in prices[max(0, dc_end - 252):dc_end + 1])
    dd_end = (prices[dc_end][1] / y1_high - 1) * 100
    low_in_dc = min(p[1] for p in prices[dc_start:dc_end + 1])
    ma20_end = ma_at(dc_end, 20)
    above_ma20_end = prices[dc_end][1] > ma20_end if ma20_end else False
    print(f"\n  ** ONGOING DEATH CROSS ** {prices[dc_start][0][:10]} ~ {prices[dc_end][0][:10]} "
          f"({dc_days}d) DD={dd_end:+.1f}% MA20={'ABOVE' if above_ma20_end else 'below'} "
          f"Low={low_in_dc:,.0f}")


# ============================================
# Part 2: Classification by macro regime
# ============================================
print()
print("=" * 115)
print("  Death Cross Classification by Macro Regime")
print("=" * 115)

for dp in dc_periods:
    year = dp['start'][:4]
    # Rough regime classification
    if year in ['2008']:
        regime = "GFC panic"
    elif year in ['2009', '2010']:
        regime = "QE era"
    elif year in ['2013', '2015', '2016', '2018']:
        regime = "Fed taper/hike"
    elif year in ['2019', '2020']:
        regime = "Fed cut/QE"
    elif year in ['2022']:
        regime = "Fed mega-hike"
    elif year in ['2024', '2025', '2026']:
        regime = "Fed cut + de-dollarization"
    else:
        regime = "Transition"

    outcome = "Bull resumed" if dp['fwd'][6] > 10 else ("Bear" if dp['fwd'][6] < -5 else "Sideways")
    print(f"  {dp['start'][:10]}~{dp['end'][:10]} | {regime:<25} | DD={dp['dd_end']:>+5.1f}% | 6M: {dp['fwd'][6]:>+5.1f}% | {outcome}")


# ============================================
# Part 3: MA20 cross performance during DC
# ============================================
print()
print("=" * 115)
print("  MA20 Signal Performance DURING Death Cross (buy when MA20 crossed above)")
print("=" * 115)

for dp in dc_periods:
    start_idx = next(i for i, p in enumerate(prices) if p[0] == dp['start'])
    end_idx = next(i for i, p in enumerate(prices) if p[0] == dp['end'])
    trades = []
    in_trade = False
    entry_price = 0
    for idx in range(start_idx, end_idx + 1):
        ma20 = ma_at(idx, 20)
        if ma20 is None:
            continue
        above = prices[idx][1] > ma20
        if above and not in_trade:
            entry_price = prices[idx][1]
            in_trade = True
        elif not above and in_trade:
            ret = (prices[idx][1] / entry_price - 1) * 100
            trades.append(round(ret, 1))
            in_trade = False
    if in_trade:
        ret = (prices[end_idx][1] / entry_price - 1) * 100
        trades.append(round(ret, 1))
    if trades:
        avg_ret = sum(trades) / len(trades)
        win_pct = sum(1 for t in trades if t > 0) / len(trades) * 100
        print(f"  {dp['start'][:10]}~{dp['end'][:10]}: {len(trades)} trades, avg={avg_ret:+.1f}%, win={win_pct:.0f}%  [{', '.join(f'{t:+.1f}%' for t in trades)}]")
    else:
        print(f"  {dp['start'][:10]}~{dp['end'][:10]}: no MA20 crosses")


# ============================================
# Part 4: The 2022 Analog - deep dive
# ============================================
print()
print("=" * 115)
print("  2022 DC vs 2026 DC: The Only Fed-Easing Death Crosses")
print("=" * 115)

# 2022 DC was during Fed tightening. 2026 is during Fed easing.
# This is a critical difference. Let's highlight it.

fed_tightening_dc = []
fed_easing_dc = []
for dp in dc_periods:
    year = dp['start'][:4]
    if year in ['2008', '2013', '2015', '2016', '2018', '2022']:
        fed_tightening_dc.append(dp)
    elif year in ['2009', '2010', '2019', '2020', '2024', '2025', '2026']:
        fed_easing_dc.append(dp)

print()
print("  Fed TIGHTENING Death Crosses (fundamentals bearish + technical bearish):")
for dp in fed_tightening_dc:
    print(f"    {dp['start'][:10]}~{dp['end'][:10]}: DD={dp['dd_end']:+.1f}%, 6M_fwd={dp['fwd'][6]:+.1f}%")

print()
print("  Fed EASING Death Crosses (fundamentals bullish + technical bearish) -- current regime:")
for dp in fed_easing_dc:
    marker = " <-- NOW" if dp['end'] >= '2026-07' else ""
    print(f"    {dp['start'][:10]}~{dp['end'][:10]}: DD={dp['dd_end']:+.1f}%, 6M_fwd={dp['fwd'][6]:+.1f}%{marker}")

if fed_easing_dc:
    completed = [dp for dp in fed_easing_dc if dp['end'] < '2026-07']
    if completed:
        avg_6m = sum(dp['fwd'][6] for dp in completed) / len(completed)
        all_positive = all(dp['fwd'][6] > 0 for dp in completed)
        print(f"\n  Easing DCs (completed, n={len(completed)}): avg 6M forward = {avg_6m:+.1f}%, all positive = {all_positive}")

# ============================================
# Part 5: Summary
# ============================================
print()
print("=" * 115)
print("  SUMMARY")
print("=" * 115)
print(f"  Total Death Cross periods 2008-2026: {len(dc_periods)}")

for m in [1, 3, 6, 12]:
    rets = [dp['fwd'][m] for dp in dc_periods]
    avg = sum(rets) / len(rets)
    pos = sum(1 for r in rets if r > 0) / len(rets) * 100
    print(f"  {m}M after DC ends: avg {avg:+.1f}%, positive {pos:.0f}% of time")

ma20_above = [dp for dp in dc_periods if dp['above_ma20']]
ma20_below = [dp for dp in dc_periods if not dp['above_ma20']]
if ma20_above:
    print(f"\n  DC ending with MA20 ABOVE ({len(ma20_above)}/{len(dc_periods)}): avg 6M = {sum(dp['fwd'][6] for dp in ma20_above)/len(ma20_above):+.1f}%")
if ma20_below:
    print(f"  DC ending with MA20 BELOW ({len(ma20_below)}/{len(dc_periods)}): avg 6M = {sum(dp['fwd'][6] for dp in ma20_below)/len(ma20_below):+.1f}%")

print()
print(f"  Current DC: {prices[dc_start][0][:10]} ~ ongoing ({n-1-dc_start+1}d)")
print(f"  Price: {prices[n-1][1]:.0f}, MA20: {ma_at(n-1,20):.0f}, MA50: {ma_at(n-1,50):.0f}, MA200: {ma_at(n-1,200):.0f}")
print(f"  DD from 1Y high: {(prices[n-1][1]/max(p[1] for p in prices[-252:])-1)*100:+.1f}%")
